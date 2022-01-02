from openpyxl import load_workbook, Workbook

# 打开【test.xlsx】工作簿
wb = load_workbook('./test.xlsx')
# 获取活动工作表
ws = wb.active

# 获取表头
header = []
for cell in ws[1]:
    header.append(cell.value)

# 新建工作簿
new_wb = Workbook()
# 获取新工作簿中的工作表
new_ws = new_wb.active

# 将表头写入新工作簿的工作表中
new_ws.append(header)

# 从第二行开始遍历表格
for row in ws.iter_rows(min_row=2, values_only=True):
    # 取出姓名，迟到时间和迟到次数
    name = row[1]
    time = row[3]
    number = row[-1]
    # 判断是否迟到
    if time > 45 and number > 3:
        print('{}迟到了{}分钟，迟到了{}次'.format(name, time, number))
        # 将迟到人员信息写入新工作簿的工作表中
        new_ws.append(row)

# 将新工作簿保存为【new.xlsx】
new_wb.save('./new.xlsx')